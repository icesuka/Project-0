from selenium.webdriver import Chrome
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from openpyxl import Workbook
import time

opt=Options()
opt.add_argument('--disable-blink-features=AutomationControlled')
web=Chrome(options=opt)
wb=Workbook()
ws1=wb.active
ws1.title='队伍榜'
ws2=wb.create_sheet('选手榜')
ws3=wb.create_sheet('英雄榜')

web.get('https://www.fnscore.cn/statistics.html')
web.maximize_window()
time.sleep(1)

ele1=web.find_element(By.XPATH,'//*[@id="__layout"]/div/div[2]/div[1]/div[2]/div/div')
ActionChains(web).move_to_element(ele1).perform()
js1='window.scrollTo(0,300);'
web.execute_script(js1)
time.sleep(0.5)
web.find_element(By.XPATH,'//*[@id="__layout"]/div/div[2]/div[1]/div[2]/div/ul/li[26]/span').click()
time.sleep(0.5)
js2='window.scrollTo(0,0);'
web.execute_script(js2)
time.sleep(0.5)

for i in range(18):
    dataframe1=[]
    for j in range(17):
        if i == 0:
            data1=web.find_element(By.XPATH,f'//*[@id="__layout"]/div/div[2]/div[3]/div[1]/table/tr[{i+1}]/th[{j+1}]/p').text
        else:
            try:
                data1=web.find_element(By.XPATH,f'//*[@id="__layout"]/div/div[2]/div[3]/div[1]/table/tr[{i+1}]/td[{j+1}]/p').text
            except:
                data1=web.find_element(By.XPATH,f'//*[@id="__layout"]/div/div[2]/div[3]/div[1]/table/tr[{i+1}]/td[{j+1}]/a').text
        dataframe1.append(data1)
    print(dataframe1)
    ws1.append(dataframe1)
time.sleep(2)

ele2=web.find_element(By.XPATH,'//*[@id="__layout"]/div/div[2]/div[1]/div[2]/div/div')
ActionChains(web).move_to_element(ele2).perform()
web.execute_script(js1)
time.sleep(0.5)
web.find_element(By.XPATH,'//*[@id="__layout"]/div/div[2]/div[1]/div[2]/div/ul/li[26]/span').click()
time.sleep(0.5)
web.execute_script(js2)
time.sleep(0.5)
web.find_element(By.XPATH,'//*[@id="__layout"]/div/div[2]/div[2]/div[2]/p').click()
time.sleep(0.5)

for m in range(119):
    dataframe2=[]
    for n in range(13):
        if m == 0:
            data2=web.find_element(By.XPATH,f'//*[@id="__layout"]/div/div[2]/div[3]/div[1]/table/tr[{m+1}]/th[{n+1}]/p').text
        else:
            try:
                data2=web.find_element(By.XPATH,f'//*[@id="__layout"]/div/div[2]/div[3]/div[1]/table/tr[{m+1}]/td[{n+1}]/p').text
            except:
                data2=web.find_element(By.XPATH,f'//*[@id="__layout"]/div/div[2]/div[3]/div[1]/table/tr[{m+1}]/td[{n+1}]/a').text
        dataframe2.append(data2)
    print(dataframe2)
    ws2.append(dataframe2)
time.sleep(1)

ele3=web.find_element(By.XPATH,'//*[@id="__layout"]/div/div[2]/div[1]/div[2]/div/div')
ActionChains(web).move_to_element(ele3).perform()
web.execute_script(js1)
time.sleep(0.5)
web.find_element(By.XPATH,'//*[@id="__layout"]/div/div[2]/div[1]/div[2]/div/ul/li[26]/span').click()
time.sleep(0.5)
web.execute_script(js2)
time.sleep(0.5)
ele4=web.find_element(By.XPATH,'//*[@id="__layout"]/div/div[2]/div[2]/div[1]/p')
ActionChains(web).move_to_element(ele4).perform()
time.sleep(0.5)
web.find_element(By.XPATH,'//*[@id="__layout"]/div/div[2]/div[2]/div[3]/p').click()
time.sleep(0.5)

for x in range(107):
    dataframe3=[]
    for y in range(6):
        if x == 0:
            data3=web.find_element(By.XPATH,f'//*[@id="__layout"]/div/div[2]/div[3]/div[1]/table/tr[{x+1}]/th[{y+1}]/p').text
        else:
            try:
                data3=web.find_element(By.XPATH,f'//*[@id="__layout"]/div/div[2]/div[3]/div[1]/table/tr[{x+1}]/td[{y+1}]/p').text
            except:
                data3=web.find_element(By.XPATH,f'//*[@id="__layout"]/div/div[2]/div[3]/div[1]/table/tr[{x+1}]/td[{y+1}]/a').text
        dataframe3.append(data3)
    print(dataframe3)
    ws3.append(dataframe3)
wb.save('蜂鸟电竞-2022LPL春季赛常规赛.xlsx')
time.sleep(2)
web.quit()
