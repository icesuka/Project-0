from selenium.webdriver import Chrome
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from openpyxl import Workbook
import time

opt=Options()
opt.add_argument('--disable-blink-features=AutomationControlled')
web=Chrome(options=opt)
wb=Workbook()
ws1=wb.active
ws1.title='队伍榜'
ws2=wb.create_sheet('选手榜')
ws3=wb.create_sheet('英雄榜')

web.get('https://lpl.qq.com/esnew/data/rank.shtml?iGameId=167&sGameType=1,5')
web.maximize_window()
js1='window.scrollTo(0,300);'
js2='window.scrollTo(0,000);'
time.sleep(1)

web.execute_script(js1)
time.sleep(0.5)
for i in range(18):
    dataframe1=[]
    for j in range(12):
        if i == 0:
            data1=web.find_element(By.XPATH,f'//*[@id="data-page"]/div[5]/div[1]/div[2]/div/table/thead/tr/th[{j+1}]').text
        else:
            try:
                data1 = web.find_element(By.XPATH,f'[@id="teamRank"]/tr[{i}]/td[{j+1}]/span').text
            except:
                data1 = web.find_element(By.XPATH,f'//*[@id="teamRank"]/tr[{i}]/td[{j+1}]').text
        dataframe1.append(data1)
    print(dataframe1)
    ws1.append(dataframe1)
web.execute_script(js2)
time.sleep(2)

web.find_element(By.XPATH,'//*[@id="data-page"]/div[4]/ul/li[2]/a').click()
web.execute_script(js1)
time.sleep(0.5)
for m in range(114):
    dataframe2=[]
    for n in range(17):
        if m == 0:
            data2=web.find_element(By.XPATH,f'//*[@id="data-page"]/div[5]/div[2]/div[2]/div/table/thead/tr/th[{n+1}]').text
        else:
            try:
                data2 = web.find_element(By.XPATH,f'//[@id="playerRank"]/tr[{m}]/td[{n+1}]/span').text
            except:
                data2 = web.find_element(By.XPATH,f'//*[@id="playerRank"]/tr[{m}]/td[{n+1}]').text
        dataframe2.append(data2)
    print(dataframe2)
    ws2.append(dataframe2)
web.execute_script(js2)
time.sleep(2)

web.find_element(By.XPATH,'//*[@id="data-page"]/div[4]/ul/li[3]/a').click()
web.execute_script(js1)
time.sleep(0.5)
for x in range(105):
    dataframe3=[]
    for y in range(11):
        if x == 0:
            data3=web.find_element(By.XPATH,f'//*[@id="data-page"]/div[5]/div[3]/div[2]/div/table/thead/tr/th[{y+1}]').text
        else:
            try:
                data3 = web.find_element(By.XPATH,f'//[@id="heroRank"]/tr[{x}]/td[{y+1}]/span').text
            except:
                data3 = web.find_element(By.XPATH,f'//*[@id="heroRank"]/tr[{x}]/td[{y+1}]').text
        dataframe3.append(data3)
    print(dataframe3)
    ws3.append(dataframe3)
wb.save('LPL官网-2022LPL春季赛常规赛.xlsx')
time.sleep(2)
web.quit()

